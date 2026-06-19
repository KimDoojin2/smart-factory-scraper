"""
스마트공장 공급기업 수집기 — SmartFind v2
체크박스 필터 → 수집 → 엑셀 다운로드 → 드래그앤드롭 업데이트
"""
from __future__ import annotations

import io, json, time, random
from datetime import datetime
from pathlib import Path

import streamlit as st
import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "output"
DATA_DIR.mkdir(exist_ok=True)

BASE_URL = "https://www.smart-factory.kr"
API_URL = f"{BASE_URL}/usr/bg/fs/ma/FixesSplySrch/selectFixesSplyContainer.do"
MAX_PAGES = 250

st.set_page_config(page_title="SmartFind", page_icon="🏭", layout="wide")

# ── 전문분야 체크박스 구조 (스크린샷 기준) ──
SPECIALTY_FIELDS = {
    "자동화기기": ["산업용로봇", "스마트장비", "식별시스템/머신비전"],
    "연결화기기": ["제어시스템/컨트롤러", "센서/액추에이터", "통신네트워크장비"],
    "정보화솔루션": ["ERP", "PLM/PDM", "SCM/WMS", "MES/POP/APS", "FEMS", "QMS/CMMS"],
    "지능형서비스": ["클라우드컴퓨팅", "CPS/디지털트윈", "제조빅데이터/제조AI", "AR/VR/MR", "컨설팅/사이버보안"],
}

INDUSTRY_TYPES = [
    "식료품 제조업", "음료 제조업", "담배 제조업",
    "섬유제품 제조업; 의복 제외", "의복, 의복 액세서리 및 모피제품 제조업",
    "가죽, 가방 및 신발 제조업",
    "목재 및 나무제품 제조업; 가구 제외",
    "펄프, 종이 및 종이제품 제조업", "인쇄 및 기록매체 복제업",
    "코크스, 연탄 및 석유정제품 제조업",
    "화학 물질 및 화학제품 제조업; 의약품 제외",
    "의료용 물질 및 의약품 제조업",
    "고무 및 플라스틱제품 제조업", "비금속 광물제품 제조업",
    "1차 금속 제조업",
    "금속 가공제품 제조업; 기계 및 가구 제외",
    "전자 부품, 컴퓨터, 영상, 음향 및 통신장비 제조업",
    "의료, 정밀, 광학 기기 및 시계 제조업",
    "전기장비 제조업", "기타 기계 및 장비 제조업",
    "자동차 및 트레일러 제조업", "기타 운송장비 제조업",
    "가구 제조업", "기타 제품 제조업",
    "산업용 기계 및 장비 수리업",
]

# ── CSS ──
st.markdown("""<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');
.stApp{font-family:'Inter',system-ui,sans-serif}
#MainMenu,footer{visibility:hidden}
header[data-testid="stHeader"]{background:transparent!important}
.block-container{padding-top:1rem!important}

.hero{text-align:center;padding:24px 0 16px}
.hero .brand{font-size:2rem;font-weight:900;letter-spacing:-2px}
.hero .brand span{color:#111}.hero .brand b{color:#0066FF}
.hero .sub{color:#9CA3AF;font-size:.78rem;font-weight:500;letter-spacing:2px}

.stat-row{display:flex;gap:10px;justify-content:center;margin:16px 0}
.stat-box{background:#F8FAFF;border:1px solid #E5E9F2;border-radius:10px;padding:12px 20px;text-align:center}
.stat-box .num{font-size:1.3rem;font-weight:800;color:#111;letter-spacing:-1px}
.stat-box .lbl{font-size:.62rem;color:#9CA3AF;font-weight:600;text-transform:uppercase;letter-spacing:1px;margin-top:2px}

.filter-title{font-size:.82rem;font-weight:700;color:#1B3A5C;margin:8px 0 4px;padding-left:2px}
.filter-cat{font-size:.72rem;font-weight:600;color:#6B7280;background:#F1F5F9;padding:4px 10px;border-radius:6px;display:inline-block;margin:4px 0 2px}

.done-box{background:linear-gradient(135deg,#F0FDF4,#ECFDF5);border:1px solid #BBF7D0;border-radius:14px;padding:24px;text-align:center;margin:16px 0}
.done-box .icon{font-size:2rem;margin-bottom:4px}
.done-box .msg{font-size:1rem;font-weight:700;color:#059669}
.done-box .detail{font-size:.78rem;color:#6B7280;margin-top:4px}
</style>""", unsafe_allow_html=True)


# ── API ──
def _fetch_all(progress_bar, status_text) -> list[dict]:
    s = requests.Session()
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/json", "Content-Type": "application/json",
        "Referer": f"{BASE_URL}/usr/bg/fs/ma/fixesSplySrch", "Origin": BASE_URL,
    })
    s.get(BASE_URL, timeout=15)

    first = s.post(API_URL, json={"key": "splyList"}, timeout=30).json()
    total_pages = min(int(first["paginationInfo"]["totalPageCount"]), MAX_PAGES)
    total_count = int(first["paginationInfo"]["totalCount"])
    all_items = list(first.get("splyInstList", []))
    status_text.text(f"전체 {total_count:,}건 · {total_pages}페이지 수집 중...")

    for page in range(2, total_pages + 1):
        time.sleep(random.uniform(0.15, 0.35))
        payload = {"key": "splyList", "currentPage": str(page), "recordCountPerPage": "50"}
        for attempt in range(3):
            try:
                items = s.post(API_URL, json=payload, timeout=30).json().get("splyInstList", [])
                all_items.extend(items)
                break
            except Exception:
                time.sleep(2 ** attempt)
        progress_bar.progress(page / total_pages, text=f"{page}/{total_pages} 페이지 · {len(all_items):,}건")

    progress_bar.progress(1.0, text="완료!")
    return all_items


def _apply_filters(items: list[dict], tab: str, specialties: list[str], industries: list[str]) -> list[dict]:
    filtered = items

    if tab != "전체":
        filtered = [i for i in filtered if tab in str(i.get("splyInstSe", ""))]

    if specialties:
        def match_specialty(item):
            fields = str(item.get("splyMfrcSpcltyFldNm", "")) + "," + str(item.get("splySpcltyFldNm", ""))
            return any(sp in fields for sp in specialties)
        filtered = [i for i in filtered if match_specialty(i)]

    if industries:
        def match_industry(item):
            biz = str(item.get("splyTpbizNm", ""))
            return any(ind in biz for ind in industries)
        filtered = [i for i in filtered if match_industry(i)]

    return filtered


# ── Excel ──
def _build_excel(items: list[dict], filter_desc: str = "") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "공급기업 목록"

    columns = [
        ("instNm",              "업체명",       24),
        ("splyInstSe",          "공급기업 구분", 22),
        ("splyCnstcNocs",       "구축건수",     10),
        ("splyLctnCdNm",        "지역",         8),
        ("splyWholEmpCnt",      "종사자규모",   12),
        ("splySlsAmt",          "매출규모(억)", 14),
        ("slsYr",               "매출연도",     10),
        ("splyDgstfnScr",       "만족도(5점)",  12),
        ("splyMfrcSpcltyFldNm", "주력분야",     30),
        ("splySpcltyFldNm",     "전문분야",     42),
        ("splyTpbizNm",         "특화업종",     48),
        ("rprsvNm",             "대표자",       10),
        ("fndnYmd",             "설립일",       12),
        ("instAddr",            "주소",         38),
        ("rprsTelno",           "대표전화",     16),
        ("rprsFxno",            "팩스",         16),
        ("hmpgAddr",            "홈페이지",     26),
        ("brno",                "사업자번호",   14),
    ]

    keys = [c[0] for c in columns]
    headers = [c[1] for c in columns]
    widths = [c[2] for c in columns]

    hdr_font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_font = Font(name="맑은 고딕", size=9)
    cell_font_bold = Font(name="맑은 고딕", size=9, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")

    thin = Side(style="thin", color="D9DEE4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    stripe_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    green_font = Font(name="맑은 고딕", size=9, color="059669", bold=True)
    blue_font = Font(name="맑은 고딕", size=9, color="2563EB")
    orange_font = Font(name="맑은 고딕", size=9, color="D97706", bold=True)

    ws.row_dimensions[1].height = 32
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border

    for row_idx, item in enumerate(items, 2):
        is_stripe = row_idx % 2 == 0
        fill = stripe_fill if is_stripe else white_fill
        ws.row_dimensions[row_idx].height = 22

        for col_idx, key in enumerate(keys, 1):
            val = item.get(key, "")
            if val is None or str(val).strip() in ("None", "null", "."):
                val = ""

            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.fill = fill

            if key in ("splyCnstcNocs", "splyWholEmpCnt"):
                try:
                    val = int(val) if val else 0
                except (ValueError, TypeError):
                    val = 0
                cell.value = val
                cell.alignment = center
                cell.font = green_font if val > 0 else cell_font

            elif key == "splySlsAmt":
                try:
                    val = int(val) if val else 0
                except (ValueError, TypeError):
                    val = 0
                cell.value = f"{val}억" if val else "-"
                cell.alignment = center
                cell.font = cell_font_bold if val else cell_font

            elif key == "splyDgstfnScr":
                try:
                    val = float(val) if val else 0
                except (ValueError, TypeError):
                    val = 0
                if val > 0:
                    stars = "★" * int(val) + ("½" if val % 1 >= 0.5 else "")
                    cell.value = f"{stars} ({val})"
                    cell.font = orange_font
                else:
                    cell.value = "-"
                    cell.font = cell_font
                cell.alignment = center

            elif key == "instNm":
                cell.value = str(val)
                cell.font = cell_font_bold
                cell.alignment = left

            elif key == "splyInstSe":
                cell.value = str(val)
                cell.alignment = center
                if "제조" in str(val) and "서비스" in str(val):
                    cell.font = Font(name="맑은 고딕", size=9, color="7C3AED", bold=True)
                elif "서비스" in str(val):
                    cell.font = Font(name="맑은 고딕", size=9, color="0891B2", bold=True)
                elif "공방" in str(val):
                    cell.font = Font(name="맑은 고딕", size=9, color="D97706", bold=True)
                else:
                    cell.font = blue_font

            elif key in ("splyMfrcSpcltyFldNm", "splySpcltyFldNm", "splyTpbizNm", "instAddr"):
                cell.value = str(val) if val else "-"
                cell.font = cell_font
                cell.alignment = left

            elif key == "hmpgAddr":
                cell.value = str(val) if val else ""
                cell.font = Font(name="맑은 고딕", size=9, color="2563EB", underline="single") if val else cell_font
                cell.alignment = left

            else:
                cell.value = str(val) if val else "-"
                cell.font = cell_font
                cell.alignment = center

    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # ── 요약 시트 ──
    ws2 = wb.create_sheet("요약")
    ws2.sheet_properties.tabColor = "0066FF"
    summary_hdr_fill = PatternFill(start_color="0066FF", end_color="0066FF", fill_type="solid")
    summary_hdr_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    summary_font = Font(name="맑은 고딕", size=10)
    summary_val_font = Font(name="맑은 고딕", size=12, bold=True, color="1B3A5C")

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 20

    mfg = sum(1 for i in items if "제조" in str(i.get("splyInstSe", "")))
    svc = sum(1 for i in items if "서비스" in str(i.get("splyInstSe", "")))
    wshop = sum(1 for i in items if "공방" in str(i.get("splyInstSe", "")))

    summary_data = [
        ("수집 결과", f"{len(items):,}건"),
        ("스마트제조", f"{mfg:,}건"),
        ("스마트서비스", f"{svc:,}건"),
        ("스마트공방", f"{wshop:,}건"),
        ("수집일시", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("필터 조건", filter_desc or "전체"),
        ("데이터 출처", "smart-factory.kr"),
    ]

    for idx, lbl in enumerate(["항목", "값"], 1):
        c = ws2.cell(row=1, column=idx, value=lbl)
        c.font = summary_hdr_font
        c.fill = summary_hdr_fill
        c.alignment = center
        c.border = border

    for r, (lbl, val) in enumerate(summary_data, 2):
        c1 = ws2.cell(row=r, column=1, value=lbl)
        c1.font = summary_font
        c1.border = border
        c2 = ws2.cell(row=r, column=2, value=val)
        c2.font = summary_val_font
        c2.border = border
        c2.alignment = center

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


EXCEL_HEADERS = [c[1] for c in [
    ("instNm", "업체명", 24), ("splyInstSe", "공급기업 구분", 22),
    ("splyCnstcNocs", "구축건수", 10), ("splyLctnCdNm", "지역", 8),
    ("splyWholEmpCnt", "종사자규모", 12), ("splySlsAmt", "매출규모(억)", 14),
    ("slsYr", "매출연도", 10), ("splyDgstfnScr", "만족도(5점)", 12),
    ("splyMfrcSpcltyFldNm", "주력분야", 30), ("splySpcltyFldNm", "전문분야", 42),
    ("splyTpbizNm", "특화업종", 48), ("rprsvNm", "대표자", 10),
    ("fndnYmd", "설립일", 12), ("instAddr", "주소", 38),
    ("rprsTelno", "대표전화", 16), ("rprsFxno", "팩스", 16),
    ("hmpgAddr", "홈페이지", 26), ("brno", "사업자번호", 14),
]]

EXCEL_KEYS = [
    "instNm", "splyInstSe", "splyCnstcNocs", "splyLctnCdNm",
    "splyWholEmpCnt", "splySlsAmt", "slsYr", "splyDgstfnScr",
    "splyMfrcSpcltyFldNm", "splySpcltyFldNm", "splyTpbizNm",
    "rprsvNm", "fndnYmd", "instAddr", "rprsTelno", "rprsFxno",
    "hmpgAddr", "brno",
]


def _parse_uploaded_excel(uploaded_file) -> list[dict]:
    """업로드된 엑셀에서 기존 데이터를 dict 리스트로 변환."""
    import re
    df = pd.read_excel(uploaded_file, sheet_name=0)
    header_to_key = dict(zip(EXCEL_HEADERS, EXCEL_KEYS))
    df = df.rename(columns={h: header_to_key.get(h, h) for h in df.columns})
    records = []
    for _, row in df.iterrows():
        rec = {}
        for key in EXCEL_KEYS:
            val = row.get(key, "")
            if pd.isna(val):
                val = ""
            val = str(val).strip()
            # 만족도 "★★★ (3.5)" → 3.5
            if key == "splyDgstfnScr" and "★" in val:
                m = re.search(r"\(([\d.]+)\)", val)
                val = m.group(1) if m else "0"
            # 매출 "32억" → 32
            if key == "splySlsAmt" and "억" in val:
                val = val.replace("억", "").strip()
            rec[key] = val
        records.append(rec)
    return records


def _merge_data(old_items: list[dict], new_items: list[dict]) -> tuple[list[dict], int, int]:
    """기존 + 신규 병합. instCd(기관코드) 또는 업체명+사업자번호로 중복 판별.
    Returns: (merged, new_count, updated_count)
    """
    def _item_key(item):
        code = str(item.get("instCd", "") or item.get("기관코드", "")).strip()
        if code:
            return code
        name = str(item.get("instNm", "") or item.get("업체명", "")).strip()
        brno = str(item.get("brno", "") or item.get("사업자번호", "")).strip()
        return f"{name}|{brno}"

    old_map = {}
    for item in old_items:
        k = _item_key(item)
        if k:
            old_map[k] = item

    merged = list(old_items)
    new_count = 0
    updated_count = 0

    for item in new_items:
        k = _item_key(item)
        if not k:
            continue

        if k in old_map:
            old = old_map[k]
            changed = False
            for field in EXCEL_KEYS:
                new_val = str(item.get(field, "") or "").strip()
                old_val = str(old.get(field, "") or "").strip()
                # 만족도/매출 등 숫자 비교 정규화
                if new_val and new_val not in ("-", "0", "") and new_val != old_val:
                    old[field] = item[field]
                    changed = True
            if changed:
                updated_count += 1
        else:
            merged.append(item)
            old_map[k] = item
            new_count += 1

    return merged, new_count, updated_count


# =============================================================================
#  UI
# =============================================================================
st.markdown("""<div class="hero">
<div class="brand"><span>SMART</span><b>FIND</b></div>
<div class="sub">스마트공장 공급기업 수집기</div>
</div>""", unsafe_allow_html=True)

# ── 데이터 로드 ──
if "all_data" not in st.session_state:
    latest = sorted(DATA_DIR.glob("smart_factory_suppliers_*.json"), reverse=True)
    if latest:
        with open(latest[0], "r", encoding="utf-8") as f:
            st.session_state["all_data"] = json.load(f)
        st.session_state["last_ts"] = latest[0].stem.replace("smart_factory_suppliers_", "")
    else:
        st.session_state["all_data"] = []
        st.session_state["last_ts"] = ""

all_data = st.session_state["all_data"]

# ── 현황 표시 ──
if all_data:
    total = len(all_data)
    mfg = sum(1 for i in all_data if "제조" in str(i.get("splyInstSe", "")))
    svc = sum(1 for i in all_data if "서비스" in str(i.get("splyInstSe", "")))
    wshop = sum(1 for i in all_data if "공방" in str(i.get("splyInstSe", "")))
    st.markdown(f"""<div class="stat-row">
    <div class="stat-box"><div class="num">{total:,}</div><div class="lbl">전체</div></div>
    <div class="stat-box"><div class="num">{mfg:,}</div><div class="lbl">제조</div></div>
    <div class="stat-box"><div class="num">{svc:,}</div><div class="lbl">서비스</div></div>
    <div class="stat-box"><div class="num">{wshop:,}</div><div class="lbl">공방</div></div>
    </div>""", unsafe_allow_html=True)
    if st.session_state["last_ts"]:
        st.caption(f"📅 최근 수집: {st.session_state['last_ts']}")

st.markdown("---")

# ── 수집 버튼 ──
if st.button("🚀 전체 수집 시작" if not all_data else "🔄 새로 수집", type="primary", use_container_width=True):
    progress = st.progress(0, text="연결 중...")
    status = st.empty()

    items = _fetch_all(progress, status)

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    with open(DATA_DIR / f"smart_factory_suppliers_{ts}.json", "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)

    st.session_state["all_data"] = items
    st.session_state["last_ts"] = ts

    st.markdown(f"""<div class="done-box">
    <div class="icon">✅</div>
    <div class="msg">수집 완료! {len(items):,}건</div>
    <div class="detail">{ts} · smart-factory.kr</div>
    </div>""", unsafe_allow_html=True)
    time.sleep(1.5)
    st.rerun()

if not all_data:
    st.stop()

# =============================================================================
#  필터 영역
# =============================================================================
st.markdown("---")
st.markdown("### 🔍 필터 선택")
st.caption("체크 후 아래 '엑셀 다운로드' 버튼을 누르면 해당 조건의 기업만 추출됩니다.")

# ── 탭 (구분) ──
tab_col, _ = st.columns([2, 3])
with tab_col:
    tab = st.radio("공급기업 구분", ["전체", "스마트제조", "스마트서비스", "스마트공방"], horizontal=True)

# ── 전문분야 체크박스 ──
st.markdown('<div class="filter-title">■ 스마트공장 전문, 주력분야</div>', unsafe_allow_html=True)

selected_specialties = []

for cat_name, sub_items in SPECIALTY_FIELDS.items():
    st.markdown(f'<div class="filter-cat">{cat_name}</div>', unsafe_allow_html=True)
    cols = st.columns(len(sub_items))
    for i, sub in enumerate(sub_items):
        with cols[i]:
            if st.checkbox(sub, key=f"sp_{sub}"):
                selected_specialties.append(sub)

# ── 특화업종 체크박스 ──
st.markdown('<div class="filter-title">■ 공급기업 전문, 특화업종 (한국 산업분류 코드)</div>', unsafe_allow_html=True)

selected_industries = []
cols_per_row = 4
for row_start in range(0, len(INDUSTRY_TYPES), cols_per_row):
    row_items = INDUSTRY_TYPES[row_start:row_start + cols_per_row]
    cols = st.columns(cols_per_row)
    for i, ind in enumerate(row_items):
        with cols[i]:
            if st.checkbox(ind, key=f"ind_{ind}"):
                selected_industries.append(ind)

# ── 필터 적용 & 다운로드 ──
st.markdown("---")

filtered = _apply_filters(all_data, tab, selected_specialties, selected_industries)

# 필터 설명 생성
filter_parts = []
if tab != "전체":
    filter_parts.append(f"구분: {tab}")
if selected_specialties:
    filter_parts.append(f"분야: {', '.join(selected_specialties)}")
if selected_industries:
    filter_parts.append(f"업종: {', '.join(selected_industries)}")
filter_desc = " | ".join(filter_parts) if filter_parts else "전체"

# 결과 표시
total_pages = (len(filtered) + 49) // 50
col_r1, col_r2, col_r3 = st.columns(3)
col_r1.metric("필터 결과", f"{len(filtered):,}건")
col_r2.metric("페이지 수", f"{total_pages}페이지")
col_r3.metric("필터 조건", filter_desc[:20] + ("..." if len(filter_desc) > 20 else ""))

if filtered:
    excel_bytes = _build_excel(filtered, filter_desc)
    ts = datetime.now().strftime("%Y%m%d_%H%M")

    fname_parts = ["스마트공장_공급기업"]
    if tab != "전체":
        fname_parts.append(tab)
    if selected_specialties:
        fname_parts.append("_".join(s[:4] for s in selected_specialties[:3]))
    fname_parts.append(ts)
    fname = "_".join(fname_parts) + ".xlsx"

    st.download_button(
        f"📥 엑셀 다운로드 ({len(filtered):,}건)",
        excel_bytes,
        fname,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )

    # 미리보기
    with st.expander(f"📋 미리보기 (상위 20건)"):
        preview = filtered[:20]
        df = pd.DataFrame([{
            "업체명": i.get("instNm", ""),
            "구분": i.get("splyInstSe", ""),
            "주력분야": i.get("splyMfrcSpcltyFldNm", ""),
            "전문분야": i.get("splySpcltyFldNm", ""),
            "지역": i.get("splyLctnCdNm", ""),
            "구축건수": i.get("splyCnstcNocs", 0),
            "매출(억)": i.get("splySlsAmt", 0),
            "만족도": i.get("splyDgstfnScr", 0),
        } for i in preview])
        st.dataframe(df, use_container_width=True, hide_index=True)
else:
    st.warning("선택한 조건에 해당하는 기업이 없습니다.")

# =============================================================================
#  드래그앤드롭 업데이트
# =============================================================================
st.markdown("---")
st.markdown("### 📂 기존 엑셀 업데이트")
st.caption("이전에 다운받은 엑셀을 올리면, 신규/변경된 기업만 추가해서 다시 다운로드합니다.")

uploaded = st.file_uploader(
    "이전 엑셀 파일을 드래그앤드롭 하세요",
    type=["xlsx"],
    key="excel_upload",
    help="이전에 다운받은 스마트공장 공급기업 엑셀 파일",
)

if uploaded and all_data:
    try:
        old_items = _parse_uploaded_excel(uploaded)
        merged, new_cnt, upd_cnt = _merge_data(old_items, filtered)

        # 결과 표시
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("기존 데이터", f"{len(old_items):,}건")
        c2.metric("현재 수집", f"{len(filtered):,}건")
        c3.metric("신규 추가", f"+{new_cnt}건", delta=f"+{new_cnt}" if new_cnt else None)
        c4.metric("정보 갱신", f"{upd_cnt}건", delta=f"↑{upd_cnt}" if upd_cnt else None)

        if new_cnt > 0 or upd_cnt > 0:
            st.markdown(f"""<div class="done-box">
            <div class="icon">🔄</div>
            <div class="msg">병합 완료! 총 {len(merged):,}건</div>
            <div class="detail">신규 +{new_cnt}건 · 갱신 {upd_cnt}건</div>
            </div>""", unsafe_allow_html=True)

            merge_desc = f"병합 (기존 {len(old_items):,} + 신규 {new_cnt} + 갱신 {upd_cnt})"
            merged_excel = _build_excel(merged, merge_desc)
            ts = datetime.now().strftime("%Y%m%d_%H%M")

            st.download_button(
                f"📥 병합 엑셀 다운로드 ({len(merged):,}건)",
                merged_excel,
                f"스마트공장_공급기업_병합_{ts}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )
        else:
            st.success("✅ 변경 사항 없음 — 기존 엑셀이 최신 상태입니다.")

    except Exception as e:
        st.error(f"엑셀 파일 처리 오류: {e}")
