"""
스마트공장 공급기업 스크래퍼
https://www.smart-factory.kr/usr/bg/fs/ma/fixesSplySrch

전체 공급기업 목록을 API로 수집 → Excel 저장
"""
import json
import time
import random
import logging
from datetime import datetime
from pathlib import Path

import requests

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

BASE_URL = "https://www.smart-factory.kr"
API_URL = f"{BASE_URL}/usr/bg/fs/ma/FixesSplySrch/selectFixesSplyContainer.do"
PAGE_SIZE = 50
OUTPUT_DIR = Path(__file__).parent / "output"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/json, text/plain, */*",
    "Content-Type": "application/json",
    "Referer": f"{BASE_URL}/usr/bg/fs/ma/fixesSplySrch",
    "Origin": BASE_URL,
}

FIELD_MAP = {
    "instCd": "기관코드",
    "instNm": "업체명",
    "splyInstSe": "구분",
    "splySpcltyFldNm": "전문분야",
    "splyMfrcSpcltyFldNm": "제조전문분야",
    "splyTpbizNm": "업종",
    "rprsvNm": "대표자",
    "fndnYmd": "설립일",
    "instAddr": "주소",
    "instDaddr": "상세주소",
    "splyLctnCdNm": "지역",
    "rprsTelno": "대표전화",
    "rprsFxno": "팩스",
    "hmpgAddr": "홈페이지",
    "brno": "사업자번호",
    "splyWholEmpCnt": "직원수",
    "slsAmt": "매출액",
    "slsYr": "매출연도",
    "splySlsAmt": "매출액(억)",
    "splyCnstcNocs": "구축건수",
    "splyDgstfnScr": "만족도점수",
    "splyAprvYmd": "승인일",
    "splyAbtDgnsExclcEnt": "진단제외",
    "dgnsYr": "진단연도",
}


def create_session() -> requests.Session:
    s = requests.Session()
    s.headers.update(HEADERS)
    s.get(BASE_URL, timeout=15)
    return s


def fetch_page(session: requests.Session, page: int) -> dict:
    payload = {
        "key": "splyList",
        "currentPage": str(page),
        "recordCountPerPage": str(PAGE_SIZE),
    }
    for attempt in range(3):
        try:
            r = session.post(API_URL, json=payload, timeout=30)
            r.raise_for_status()
            return r.json()
        except Exception as e:
            log.warning("페이지 %d 실패 (시도 %d/3): %s", page, attempt + 1, e)
            time.sleep(2 ** attempt)
    return {}


def scrape_all() -> list[dict]:
    session = create_session()

    first = fetch_page(session, 1)
    if not first or "paginationInfo" not in first:
        log.error("첫 페이지 조회 실패")
        return []

    total = int(first["paginationInfo"]["totalCount"])
    total_pages = int(first["paginationInfo"]["totalPageCount"])
    log.info("전체 %d건, %d페이지 (페이지당 %d건)", total, total_pages, PAGE_SIZE)

    all_items = list(first.get("splyInstList", []))
    log.info("페이지 1/%d — %d건", total_pages, len(all_items))

    for page in range(2, total_pages + 1):
        time.sleep(random.uniform(0.3, 0.8))
        data = fetch_page(session, page)
        items = data.get("splyInstList", [])
        if not items:
            log.warning("페이지 %d 빈 응답 → 중단", page)
            break
        all_items.extend(items)
        if page % 20 == 0 or page == total_pages:
            log.info("페이지 %d/%d — 누적 %d건", page, total_pages, len(all_items))

    log.info("수집 완료: 총 %d건", len(all_items))
    return all_items


def save_excel(items: list[dict], filepath: Path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "공급기업목록"

    headers = list(FIELD_MAP.values())
    raw_keys = list(FIELD_MAP.keys())

    header_font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    cell_font = Font(name="맑은 고딕", size=9)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, item in enumerate(items, 2):
        for col_idx, key in enumerate(raw_keys, 1):
            val = item.get(key, "")
            if val is None:
                val = ""
            if key == "slsAmt" and val:
                try:
                    val = int(val)
                except (ValueError, TypeError):
                    pass
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.border = thin_border

    col_widths = {
        "업체명": 25, "구분": 12, "전문분야": 30, "제조전문분야": 20,
        "업종": 40, "대표자": 10, "주소": 40, "상세주소": 20,
        "지역": 8, "대표전화": 16, "팩스": 16, "홈페이지": 25,
        "사업자번호": 14, "매출액": 15,
    }
    for col, h in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = col_widths.get(h, 12)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(filepath)
    log.info("Excel 저장 완료: %s (%d건)", filepath, len(items))


def save_json(items: list[dict], filepath: Path):
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)
    log.info("JSON 저장 완료: %s (%d건)", filepath, len(items))


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")

    items = scrape_all()
    if not items:
        log.error("수집 데이터 없음")
        return

    save_excel(items, OUTPUT_DIR / f"smart_factory_suppliers_{ts}.xlsx")
    save_json(items, OUTPUT_DIR / f"smart_factory_suppliers_{ts}.json")

    log.info("완료! 총 %d개 업체 수집", len(items))


if __name__ == "__main__":
    main()
